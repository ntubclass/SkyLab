"""Export helpers for Teacher Judge rubric workflows."""

from __future__ import annotations

import io

from app.ai.teacher_judge.schemas import RubricItem

_DETECTABLE_LABELS = {
    "auto": "✅ 可自動偵測",
    "partial": "⚠️ 部分可偵測",
    "manual": "❌ 需人工評閱",
}

_DETECTABLE_COLORS = {
    "auto": "D8F5E1",  # 綠
    "partial": "FFF3CD",  # 黃
    "manual": "FDDEDE",  # 紅
}

_CHECKED_LABELS = {
    True: "✅ 已達成",
    False: "⬜ 未達成",
}


def export_to_excel(items: list[RubricItem], summary: str = "") -> bytes:
    """Generate an .xlsx file from RubricItem list, return as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "評分表"

    header_font = Font(bold=True, size=11)
    headers = [
        "項目編號",
        "評分項目",
        "說明",
        "是否達成",
        "可偵測性",
        "自動偵測方式",
        "替代建議",
    ]
    col_widths = [10, 25, 40, 12, 18, 35, 35]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths, strict=True), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="D0D0D0")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(items, start=2):
        detectable = item.detectable or "manual"
        label = _DETECTABLE_LABELS.get(detectable, detectable)
        bg_color = _DETECTABLE_COLORS.get(detectable, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg_color)

        values = [
            item.id,
            item.title,
            item.description,
            _CHECKED_LABELS.get(bool(item.checked), "⬜ 未達成"),
            label,
            item.detection_method or "",
            item.fallback or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 40

    if summary:
        last_row = len(items) + 3
        ws.cell(row=last_row, column=1, value="備註").font = Font(bold=True)
        summary_cell = ws.cell(row=last_row, column=2, value=summary)
        ws.merge_cells(
            start_row=last_row,
            start_column=2,
            end_row=last_row,
            end_column=len(headers),
        )
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[last_row].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
